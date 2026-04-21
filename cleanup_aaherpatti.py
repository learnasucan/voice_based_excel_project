#!/usr/bin/env python3
import re
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

src = Path('/Users/apple/Documents/New project 4/aaherpatti_contributions.xlsx')
dst = Path('/Users/apple/Documents/New project 4/aaherpatti_contributions.xlsx')

wb = load_workbook(src)
ws = wb['Contributions']

rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    name_mr, name_en, amt, place_mr, place_en = r
    if name_mr in (None, '', 'TOTAL'):
        continue

    name_mr = str(name_mr).strip()
    # Secondary serial cleanup for OCR variants.
    name_mr = re.sub(r'^[0-9०-९]+\s*[\)\.]?\s*', '', name_mr)
    name_mr = re.sub(r'^[-.:]+\s*', '', name_mr)
    name_mr = re.sub(r'\s+', ' ', name_mr).strip()

    if not name_mr:
        continue

    place_mr = '' if place_mr is None else str(place_mr).strip()
    name_en = '' if name_en is None else str(name_en).strip()
    place_en = '' if place_en is None else str(place_en).strip()

    letters = len(re.findall(r'[\u0900-\u097F]', name_mr))
    digits = len(re.findall(r'[0-9०-९]', name_mr))
    amount_like = bool(re.search(r'[/\\\-=]', name_mr))

    # Filter noisy OCR rows that are unlikely to be person names.
    if letters < 3:
        continue
    if digits > 0 and digits >= letters:
        continue
    if amount_like and letters < 5:
        continue
    if ' ' not in name_mr and letters < 5:
        continue

    try:
        amount = int(amt) if amt not in (None, '') else None
    except Exception:
        amount = None

    rows.append((name_mr, name_en, amount, place_mr, place_en))

seen = set()
clean = []
for row in rows:
    key = tuple('' if v is None else str(v).strip() for v in row)
    if key in seen:
        continue
    seen.add(key)
    clean.append(row)

clean.sort(key=lambda x: (x[2] if x[2] is not None else -1), reverse=True)

out = Workbook()
ows = out.active
ows.title = 'Contributions'
headers = [
    'Name (Marathi)',
    'Name (English transliteration)',
    'Contribution Amount',
    'Place (Marathi)',
    'Place (English transliteration)',
]
ows.append(headers)
for c in ows[1]:
    c.font = Font(bold=True)

total = 0
for row in clean:
    ows.append(row)
    if isinstance(row[2], int):
        total += row[2]

ows.append(['TOTAL', '', total, '', ''])
ows[f'A{ows.max_row}'].font = Font(bold=True)
ows[f'C{ows.max_row}'].font = Font(bold=True)

widths = {'A': 34, 'B': 34, 'C': 18, 'D': 22, 'E': 22}
for col, w in widths.items():
    ows.column_dimensions[col].width = w

out.save(dst)

print(f'Clean rows: {len(clean)}')
print(f'Total sum: {total}')
print('Preview:')
for row in clean[:10]:
    print(' | '.join('' if v is None else str(v) for v in row))
