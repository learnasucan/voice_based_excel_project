"""XLSX exporter."""

from __future__ import annotations

from io import BytesIO
from typing import Sequence

from ..constants import EXPORT_HEADERS
from ..types import ExportPayload, ExportRow
from .base import build_total_row, compute_total

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ModuleNotFoundError:  # pragma: no cover - handled at runtime
    Workbook = None
    Font = None


class XLSXRecordExporter:
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    extension = "xlsx"

    def export(self, rows: Sequence[ExportRow]) -> ExportPayload:
        if Workbook is None or Font is None:
            raise RuntimeError("XLSX export requires 'openpyxl'. Install it with: pip install openpyxl")

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Contributions"

        sheet.append(EXPORT_HEADERS)
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for row in rows:
            sheet.append(row.as_list())

        total = compute_total(rows)
        sheet.append(build_total_row(total))
        sheet[f"A{sheet.max_row}"].font = Font(bold=True)
        sheet[f"D{sheet.max_row}"].font = Font(bold=True)

        sheet.freeze_panes = "A2"
        widths = {
            "A": 14,
            "B": 26,
            "C": 24,
            "D": 22,
            "E": 20,
            "F": 20,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width

        buffer = BytesIO()
        workbook.save(buffer)
        return ExportPayload(data=buffer.getvalue(), total_amount=total)
