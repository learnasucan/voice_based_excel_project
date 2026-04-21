#!/usr/bin/env python3
import argparse
import os
import re
import subprocess
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from paddleocr import PaddleOCR
from indic_transliteration import sanscript
from indic_transliteration.sanscript import transliterate
from openpyxl import Workbook
from openpyxl.styles import Font


DEV_TO_ASCII = str.maketrans("०१२३४५६७८९", "0123456789")


@dataclass
class Segment:
    text: str
    score: float
    x1: int
    y1: int
    x2: int
    y2: int

    @property
    def yc(self) -> float:
        return (self.y1 + self.y2) / 2.0


@dataclass
class Record:
    name_mr: str
    amount: Optional[int]
    place_mr: str


def run(cmd: List[str]) -> None:
    subprocess.run(cmd, check=True)


def preprocess_pages(pdf_path: Path, out_dir: Path) -> List[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    raw_dir = out_dir / "raw_pages"
    raw_dir.mkdir(parents=True, exist_ok=True)

    # Convert PDF pages to PNG once.
    run(
        [
            "gs",
            "-dSAFER",
            "-dBATCH",
            "-dNOPAUSE",
            "-sDEVICE=png16m",
            "-r250",
            f"-sOutputFile={raw_dir}/page_%03d.png",
            str(pdf_path),
        ]
    )

    processed: List[Path] = []
    for raw in sorted(raw_dir.glob("page_*.png")):
        out = out_dir / f"{raw.stem}_proc.jpg"
        # Crop notebook borders and normalize for OCR stability.
        run(
            [
                "magick",
                str(raw),
                "-crop",
                "3800x5700+180+170",
                "+repage",
                "-resize",
                "1200x1800",
                "-colorspace",
                "Gray",
                "-contrast-stretch",
                "1%x1%",
                str(out),
            ]
        )
        processed.append(out)

    return processed


def normalize_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def normalize_mr_text(text: str) -> str:
    text = text.replace("|", " ").replace("\u200c", " ").replace("\u200d", " ")
    text = re.sub(r"[\[\]{}]", " ", text)
    return normalize_spaces(text)


def strip_serial_prefix(text: str) -> str:
    text = normalize_mr_text(text)
    text = re.sub(r"^[0-9०-९]+\s*[\)\.]\s*", "", text)
    text = re.sub(r"^[0-9०-९]+\s+", "", text)
    return normalize_spaces(text)


def normalize_digits(text: str) -> str:
    t = text.translate(DEV_TO_ASCII)
    # Common OCR confusions near numeric column.
    t = t.replace("O", "0").replace("o", "0").replace("D", "0")
    t = t.replace("S", "5").replace("s", "5")
    t = t.replace("x", "").replace("X", "")
    return t


def parse_amount(text: str) -> Optional[int]:
    t = normalize_digits(text)
    has_sep = bool(re.search(r"[/\\\-=]", t))
    digits = "".join(re.findall(r"\d", t))
    if not digits:
        return None

    # Handle OCR artifacts caused by '/-' suffix in handwritten amounts.
    if has_sep and digits.endswith("1"):
        if len(digits) == 3 and digits[-2] == "0":
            digits = digits[:-1] + "0"  # 201/- -> 200
        elif len(digits) >= 4:
            digits = digits[:-1]  # 2001- -> 200, 5001- -> 500

    if has_sep and digits.endswith("9") and len(digits) >= 4:
        digits = digits[:-1]  # 9009- -> 900

    # Correct leading-digit confusion for 4/5 digit strings.
    if len(digits) == 4 and int(digits) > 5000 and digits.endswith("00") and digits[0] in "893":
        cand = "2" + digits[1:]
        if int(cand) <= 5000:
            digits = cand

    if len(digits) == 5 and digits[0] in "89" and digits[2:] == "500":
        digits = "2500"  # 89500/ or 95001- often means 2500/-
    elif len(digits) >= 5 and int(digits) > 5000 and digits.endswith("00") and int(digits[1:]) <= 5000:
        digits = digits[1:]

    if int(digits) > 5000 and len(digits) >= 4 and int(digits[:-1]) <= 5000:
        digits = digits[:-1]

    val = int(digits)
    if val <= 0:
        return None

    # Keep amounts tidy (most entries are denomination-like).
    if val >= 100 and val % 50 != 0:
        val = int(((val + 25) // 50) * 50)

    while val > 5000 and val % 10 == 0:
        val //= 10
    if val > 5000 and len(str(val)) >= 4:
        tail = int(str(val)[1:])
        if tail <= 5000:
            val = tail
    return val


def looks_like_amount(seg: Segment) -> bool:
    t = normalize_digits(seg.text)
    digit_count = len(re.findall(r"\d", t))
    if seg.x1 >= 640 and digit_count >= 1:
        return True
    if digit_count >= 2 and re.search(r"[/\\-]", t):
        return True
    if seg.x1 >= 620 and digit_count >= 2:
        return True
    return False


def romanize_marathi(text: str) -> str:
    text = normalize_mr_text(text)
    if not text:
        return ""
    t = transliterate(text, sanscript.DEVANAGARI, sanscript.ITRANS)
    replace = {
        "A": "aa",
        "I": "ee",
        "U": "oo",
        "R^i": "ri",
        "R^I": "ri",
        "L^i": "li",
        "L^I": "li",
        "~N": "n",
        "~n": "n",
        "M": "m",
        "H": "h",
        "Sh": "sh",
        "S": "sh",
        "Ch": "ch",
        "JN": "gy",
    }
    for k, v in replace.items():
        t = t.replace(k, v)
    t = t.replace("^", "")
    t = re.sub(r"[^A-Za-z\s\-]", "", t)
    t = t.lower()

    words: List[str] = []
    for w in t.split():
        w = w.replace("aa", "a").replace("ee", "i").replace("oo", "u")
        w = w.replace("ii", "i").replace("uu", "u")
        w = w.replace("chh", "ch")
        if len(w) > 3 and w.endswith("a"):
            w = w[:-1]
        words.append(w.capitalize())
    return " ".join(words)


def extract_segments(result: Dict) -> List[Segment]:
    texts = result.get("rec_texts", [])
    scores = result.get("rec_scores", [])
    boxes = result.get("rec_boxes", [])
    segs: List[Segment] = []
    for text, score, box in zip(texts, scores, boxes):
        x1, y1, x2, y2 = [int(v) for v in box]
        clean = normalize_mr_text(str(text))
        if not clean:
            continue
        segs.append(Segment(clean, float(score), x1, y1, x2, y2))
    segs.sort(key=lambda s: s.yc)
    return segs


def build_records_from_segments(segs: List[Segment]) -> List[Record]:
    # Remove obvious header/footer noise.
    body = [s for s in segs if s.yc > 95 and s.yc < 1760]

    names = [s for s in body if s.x1 < 380]
    places = [s for s in body if 360 <= s.x1 < 650 and not looks_like_amount(s)]
    amounts = [s for s in body if looks_like_amount(s)]

    used_places: set[int] = set()
    used_amounts: set[int] = set()
    records: List[Record] = []

    for n_idx, name_seg in enumerate(names):
        # Drop super-short noise lines in left margin.
        if len(name_seg.text) < 3:
            continue

        # Match nearest place on same row.
        place_text = ""
        best_p: Optional[Tuple[int, float]] = None
        for p_idx, p in enumerate(places):
            if p_idx in used_places:
                continue
            dy = abs(p.yc - name_seg.yc)
            if dy <= 28:
                if best_p is None or dy < best_p[1]:
                    best_p = (p_idx, dy)
        if best_p is not None:
            used_places.add(best_p[0])
            place_text = places[best_p[0]].text

        # Match nearest amount on same row.
        amount_val: Optional[int] = None
        best_a: Optional[Tuple[int, float]] = None
        for a_idx, a in enumerate(amounts):
            if a_idx in used_amounts:
                continue
            dy = abs(a.yc - name_seg.yc)
            if dy <= 32:
                if best_a is None or dy < best_a[1]:
                    best_a = (a_idx, dy)
        if best_a is not None:
            used_amounts.add(best_a[0])
            amount_val = parse_amount(amounts[best_a[0]].text)

        name = strip_serial_prefix(name_seg.text)

        # If name line contains merged place and place column is missing, keep place blank to avoid wrong mixing.
        if name and (amount_val is not None or place_text):
            records.append(
                Record(
                    name_mr=name,
                    amount=amount_val,
                    place_mr=normalize_mr_text(place_text),
                )
            )

    return records


def dedupe_records(records: List[Record]) -> List[Record]:
    seen = set()
    out: List[Record] = []
    for r in records:
        key = (
            normalize_spaces(r.name_mr),
            r.amount if r.amount is not None else "",
            normalize_spaces(r.place_mr),
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out


def write_excel(records: List[Record], out_xlsx: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Contributions"

    headers = [
        "Name (Marathi)",
        "Name (English transliteration)",
        "Contribution Amount",
        "Place (Marathi)",
        "Place (English transliteration)",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    rows = sorted(records, key=lambda r: (r.amount if r.amount is not None else -1), reverse=True)

    total = 0
    for r in rows:
        amount = r.amount if r.amount is not None else None
        if amount is not None:
            total += amount
        ws.append(
            [
                r.name_mr,
                romanize_marathi(r.name_mr),
                amount,
                r.place_mr,
                romanize_marathi(r.place_mr),
            ]
        )

    ws.append(["TOTAL", "", total, "", ""])
    ws[f"A{ws.max_row}"].font = Font(bold=True)
    ws[f"C{ws.max_row}"].font = Font(bold=True)

    # Basic column widths for readability.
    widths = {"A": 34, "B": 34, "C": 18, "D": 22, "E": 22}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    wb.save(out_xlsx)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--pdf", required=True)
    parser.add_argument("--out", required=True)
    parser.add_argument("--workdir", default="/tmp/aaherpatti_work")
    parser.add_argument("--max-pages", type=int, default=0)
    args = parser.parse_args()

    pdf_path = Path(args.pdf).expanduser().resolve()
    out_xlsx = Path(args.out).expanduser().resolve()
    workdir = Path(args.workdir).expanduser().resolve()

    processed_pages = preprocess_pages(pdf_path, workdir)
    if args.max_pages and args.max_pages > 0:
        processed_pages = processed_pages[: args.max_pages]

    ocr = PaddleOCR(
        lang="hi",
        use_doc_orientation_classify=False,
        use_doc_unwarping=False,
        use_textline_orientation=False,
        text_det_limit_side_len=1200,
        text_recognition_batch_size=8,
    )

    all_records: List[Record] = []
    for idx, img in enumerate(processed_pages, start=1):
        result = ocr.predict(str(img))[0]
        segs = extract_segments(result)
        recs = build_records_from_segments(segs)
        all_records.extend(recs)
        print(f"Processed page {idx}/{len(processed_pages)} -> {len(recs)} rows")

    cleaned = dedupe_records(all_records)
    write_excel(cleaned, out_xlsx)

    print(f"\nTotal extracted rows (before dedupe): {len(all_records)}")
    print(f"Total rows after dedupe: {len(cleaned)}")
    print(f"Output: {out_xlsx}")

    # Preview first 10 sorted rows.
    sorted_rows = sorted(cleaned, key=lambda r: (r.amount if r.amount is not None else -1), reverse=True)
    print("\nPreview (first 10):")
    print("Name (MR) | Name (EN) | Amount | Place (MR) | Place (EN)")
    print("-" * 90)
    for r in sorted_rows[:10]:
        print(
            f"{r.name_mr} | {romanize_marathi(r.name_mr)} | {r.amount if r.amount is not None else ''} | "
            f"{r.place_mr} | {romanize_marathi(r.place_mr)}"
        )


if __name__ == "__main__":
    main()
