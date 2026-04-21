from __future__ import annotations

import csv
import importlib.util
from io import BytesIO
from unittest import TestCase

from record_export import RecordExportService
from record_export.constants import EXPORT_HEADERS

OPENPYXL_AVAILABLE = importlib.util.find_spec("openpyxl") is not None
REPORTLAB_AVAILABLE = importlib.util.find_spec("reportlab") is not None


class TestRecordExport(TestCase):
    def setUp(self) -> None:
        self.records = [
            {
                "serial_number": 1,
                "name_marathi": "राम शिंदे",
                "name_english": "Ram Shinde",
                "contribution_amount": 300,
                "place_marathi": "नाशिक",
                "place_english": "Nashik",
            },
            {
                "serial_number": 2,
                "name_marathi": "सीमा जोशी",
                "name_english": "Seema Joshi",
                "contribution_amount": 1200,
                "place_marathi": "पुणे",
                "place_english": "Pune",
            },
            {
                "serial_number": 3,
                "name_marathi": "अमोल पाटील",
                "name_english": "Amol Patil",
                "contribution_amount": 500,
                "place_marathi": "सातारा",
                "place_english": "Satara",
            },
        ]
        self.service = RecordExportService(lambda: self.records)

    def test_csv_contains_all_rows_and_total(self) -> None:
        artifact = self.service.export("csv", filename="records.csv")
        lines = artifact.payload.decode("utf-8-sig")
        parsed = list(csv.reader(lines.splitlines()))

        self.assertEqual(parsed[0], EXPORT_HEADERS)
        self.assertEqual(len(parsed), len(self.records) + 2)
        self.assertEqual(parsed[-1][0], "TOTAL")
        self.assertEqual(int(parsed[-1][3]), 2000)
        self.assertEqual(artifact.total_amount, 2000)
        self.assertEqual(artifact.row_count, 3)

    def test_xlsx_contains_headers_rows_and_total(self) -> None:
        if not OPENPYXL_AVAILABLE:
            self.skipTest("openpyxl is not installed in this environment")

        from openpyxl import load_workbook

        artifact = self.service.export("xlsx", filename="records.xlsx")
        workbook = load_workbook(filename=BytesIO(artifact.payload))
        sheet = workbook["Contributions"]

        headers = [sheet.cell(row=1, column=i).value for i in range(1, 7)]
        self.assertEqual(headers, EXPORT_HEADERS)
        self.assertEqual(sheet.max_row, len(self.records) + 2)
        self.assertEqual(sheet.cell(row=sheet.max_row, column=1).value, "TOTAL")
        self.assertEqual(sheet.cell(row=sheet.max_row, column=4).value, 2000)

    def test_pdf_export_generates_pdf_bytes(self) -> None:
        if not REPORTLAB_AVAILABLE:
            self.skipTest("reportlab is not installed in this environment")

        artifact = self.service.export("pdf", filename="records.pdf")
        self.assertTrue(artifact.payload.startswith(b"%PDF"))
        self.assertGreater(len(artifact.payload), 1000)
        self.assertEqual(artifact.total_amount, 2000)
